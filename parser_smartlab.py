import requests
import re
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl import Workbook


wb = Workbook()

ws = wb.active

ws['E1'] = 2015
ws['F1'] = 2016
ws['G1'] = 2017
ws['H1'] = 2018
ws['I1'] = 2019
ws['J1'] = 2020
ws['K1'] = 2021
ws['L1'] = 'LTM'

url = 'https://smart-lab.ru/q/shares/'
url_2 = ''
HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Safari/537.36'}
HOST = 'https://smart-lab.ru/'

def get_html(url, params=None):
	r = requests.get(url, headers=HEADERS, params=params)
	return r

def get_content(html):
	soup = BeautifulSoup(html, 'html.parser')
	items = soup.find_all('tr')
	
	global titles #делаем глобальными, т.к будем использовать их вне цикла
	titles = []
	global count
	count = 0

	for item in items:

		try:

			check = item.find('a', class_='charticon2').get('href')#делаем проверку: есть ли у данноого блока график который нам нужен, если
			#да то в подсчет нужных нам блоков добавляем один, иначе не прибавляем к количеству и переходим дальше методом continue
			if check:
				count +=1
			else:
				continue
			titles.append(HOST + item.find('a', class_='charticon2').get('href'))

		except AttributeError:
			pass

def parse():
	html = get_html(url)
	if html.status_code == 200:
		get_content(html.text)
	else:
		print('Error')



parse()


list_of_years = [2015, 2016, 2017, 2018, 2019, 2020, 2021]
list_ltm = 'LTM?'
#здесь будем поочереди переходить по нужным нам ссылкам
def get_factor(params=None):
	global fact_1
	fact_1 = []
	index_2015 = 2
	index_2016 = 2
	index_2017 = 2
	index_2018 = 2
	index_2019 = 2
	index_2020 = 2
	index_2021 = 2
	ltm_index = 2
	row_link_index = 2
	for i in range(count):
		r = requests.get(titles[i], headers=HEADERS, params=params).text
		soup = BeautifulSoup(r, 'html.parser')
		items = soup.find_all('tr', field='p_e')
		items_2 = soup.find_all('tr', class_='header_row')
		

		global nums
		nums = []
		
		
		for j in items:
			search = j.find('td', class_='chartrow').find_next_siblings('td')


			for z in items_2:
				search_2 = z.find('td', class_='chartrow').find_next_siblings('td')

			nums.append({
				'link': titles[i],
				'factors': [n.get_text(strip=True) for n in search], #генератор списка здесь нужен для перебора всех элементов в td теге, чтобы взять из них только текст, иначе не получается
				'years': [w.get_text(strip=True) for w in search_2]
				})

			
			
			

			for a, b in zip(nums[0]['years'], list_of_years):
				ws.cell(row=row_link_index, column=1, value=nums[0]['link'])
				b = str(b)
				#print(a, b)
				if a in 'LTM?':
					#print(nums[0]['factors'][-1])
					ws.cell(row=ltm_index, column=12, value=nums[0]['factors'][-1])
					ltm_index-=1
					break
					
				else:
					a = int(a)
					b = int(b)
					
					if a == b:
						#print("Совпадение:", a)
						
						if a == 2015:
							ws.cell(row=index_2015, column=5, value=nums[0]['factors'][0])
							#print("1")
							index_2015-=1
						'''
						elif a == 2016:
							ws.cell(row=index_2016, column=6, value=nums[0]['factors'][1])
							#print("2")
							index_2016-=1
						
						elif a == 2017:
							ws.cell(row=index_2017, column=7, value=nums[0]['factors'][2])
							#print("3")
							index_2017-=1

						elif a == 2018:
							ws.cell(row=index_2018, column=8, value=nums[0]['factors'][3])
							#print("4")
							index_2018-=1

						elif a == 2019:
							ws.cell(row=index_2019, column=9, value=nums[0]['factors'][4])
							#print("5")
							index_2019-=1

						elif a == 2020:
							ws.cell(row=index_2020, column=10, value=nums[0]['factors'][5])
							#print("6")
							index_2020-=1

						elif a == 2021:
							ws.cell(row=index_2021, column=11, value=nums[0]['factors'][6])
							#print("7")
							index_2021-=1'''

					else:
						continue

			row_link_index+=1


			if 'LTM?' in nums[0]['years']:
				#print("+")
				ltm_index+=2

			else:				
				ltm_index+=1
				continue


			if '2015' in nums[0]['years']:
				#print("+")
				index_2015+=2

			else:
				
				index_2015+=1
				continue
			
			'''if '2016' in nums[0]['years']:
				#print("+")
				index_2016+=2

			else:
				
				index_2016+=1
				continue

			if '2017' in nums[0]['years']:
				#print("+")
				index_2017+=2

			else:
							
				index_2017+=1
				continue

			if '2018' in nums[0]['years']:
				#print("+")
				index_2018+=2

			else:
				
				index_2018+=1
				continue

			if '2019' in nums[0]['years']:
				#print("+")
				index_2019+=2

			else:
				
				index_2019+=1
				continue

			if '2020' in nums[0]['years']:
				#print("+")
				index_2020+=2

			else:
				
				index_2020+=1
				continue

			if '2021' in nums[0]['years']:
				#print("+")
				index_2021+=2

			else:
				
				index_2021+=1
				continue'''

				#улучши свой говнокод с помощью функций


	wb.save("parser.xlsx")

	print("Save")


get_factor()


#wb.save("parser.xlsx")